from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.http import JsonResponse
from django.views.generic import ListView, CreateView, UpdateView
from django.urls import reverse_lazy
from datetime import datetime
from .forms import FormRegistro, FormTurnos, FormJustificaciones, FormTurnosProgramados, UploadTextForm, FormEmpleado
from .models import Registros, TurnosProgramados, Turnos, Justificaciones, Empleados
from openpyxl import Workbook
import json

# Basic
def home(request):
    return render(request, 'home.html')

#
def signup(request):
    if request.method == 'GET':
        return render(request, 'signup.html', {
            'form': UserCreationForm
        })

    else:
        if request.POST['password1'] == request.POST['pasword2']:
            try:
                user = User.objects.create_user(
                    request.POST['username'], request.POST['password1'])
                user.save()
                login(request, user)
                return redirect('home')
            except IntegrityError:
                render(request, 'signup.html', {
                    'form': UserCreationForm,
                    'error': 'Username already exists'
                })
    return render(request, 'signup.html', {
        'form': UserCreationForm,
        'error': 'Passwords do not match'
    })

#Silir e Ingresar sesión
def signout(request):
    logout(request)
    return redirect('home')

def signin(request):
    if request.method == 'GET':
        return render(request, 'signin.html', {
            'form': AuthenticationForm
        })
    else:
        user = authenticate(
            request, username=request.POST['username'], password=request.POST['password'])
        if user is None:
            return render(request, 'signin.html', {
                'form': AuthenticationForm,
                'error': 'Invalid username or password'
            })
        else:
            login(request, user)
            return redirect('home')

# CRU REGISTROS
@login_required
def crearRegistro(request):
    if request.method == 'GET':
        return render(request, 'crear_registro.html', {
            'form': FormRegistro
        })
    else:
        try:
            form = FormRegistro(request.POST)
            if form.is_valid():
                form.save()
            return redirect('registros')
        except ValueError:
            return render(request, 'crear_registro.html', {
                'form': FormRegistro,
                'error': 'Error al guardar el registro'
            })

@login_required
def listarRegistro(request):
    registros = Registros.objects.all()
    return render(request, 'listar_registro.html', {'registros': registros})


@login_required
def detalleRegistro(request, id_registro):
    if request.method == 'GET':
        registro = get_object_or_404(Registros, pk=id_registro)
        form = FormRegistro(instance=registro)
        return render(request, 'registro_detalle.html', {'registro': registro, 'form': form})
    else:
        try:
            registro = get_object_or_404(Registros, pk=id_registro)
            form = FormRegistro(request.POST, instance=registro)
            form.save()
            return redirect('registros')
        except ValueError:
            return render(request, 'registro_detalle.html', {'registro': registro, 'form': form, 'error': 'Error al actualizar el registro'})

# CRU Turnos Programados
@login_required
def crearProgramacion(request):
    if request.method == 'GET':
        return render(request, 'crear_programacion.html', {
            'form': FormTurnosProgramados
        })

    else:
        try:
            form = FormTurnosProgramados(request.POST)
            if form.is_valid():
                form.save()
            return redirect('programaciones')
        except ValueError:
            return render(request, 'crear_programacion.html', {
                'form': FormTurnosProgramados,
                'error': 'Error al guardar la programacion'
            })


@login_required
def listarProgramaciones(request):
    turnos_programados = TurnosProgramados.objects.all()
    return render(request, 'listar_turnos_programados.html', {'turnos_programados': turnos_programados})


@login_required
def detalleProgramacion(request, id_turno_programado):
    if request.method == 'GET':
        turno_programado = get_object_or_404(
            TurnosProgramados, pk=id_turno_programado)
        form = FormTurnos(instance=turno_programado)
        return render(request, 'turnos_programados_detalle.html', {'turno_programado': turno_programado, 'form': form})
    else:
        try:
            turno_programado = get_object_or_404(
                TurnosProgramados, pk=id_turno_programado)
            form = FormTurnos(request.POST, instance=turno_programado)
            form.save()
            return redirect('programaciones')
        except ValueError:
            return render(request, 'turnos_programados_detalle.html', {'turno_programado': turno_programado, 'form': form, 'error': 'Error al actualizar la programacion'})

# CRU Turnos
@login_required
def crearTurno(request):
    if request.method == 'GET':
        return render(request, 'crear_turno.html', {
            'form': FormTurnos
        })

    else:
        try:
            form = FormTurnos(request.POST)
            if form.is_valid():
                form.save()
            return redirect('turnos')
        except ValueError:
            return render(request, 'crear_turno.html', {
                'form': FormTurnos,
                'error': 'Error al guardar el turno'
            })


@login_required
def listarTurnos(request):
    turnos = Turnos.objects.order_by('id_turno')
    return render(request, 'listar_turnos.html', {'turnos': turnos})


@login_required
def detalleTurnos(request, id_turno):
    if request.method == 'GET':
        turno = get_object_or_404(Turnos, pk=id_turno)
        form = FormTurnos(instance=turno)
        return render(request, 'turnos_detalle.html', {'turno': turno, 'form': form})
    else:
        try:
            turno = get_object_or_404(Turnos, pk=id_turno)
            form = FormTurnos(request.POST, instance=turno)
            form.save()
            return redirect('turnos')
        except ValueError:
            return render(request, 'turnos_detalle.html', {'turno': turno, 'form': form, 'error': 'Error al actualizar la programacion'})

# CRU Justificaciones
@login_required
def crearJustificacion(request):
    if request.method == 'GET':
        return render(request, 'crear_justificacion.html', {
            'form': FormJustificaciones
        })

    else:
        try:
            form = FormJustificaciones(request.POST)
            if form.is_valid():
                form.save()
            return redirect('justificaciones')
        except ValueError:
            return render(request, 'crear_justificacion.html', {
                'form': FormJustificaciones,
                'error': 'Error al guardar la justificación'
            })

@login_required
def listarJustificaciones(request):
    justificaciones = Justificaciones.objects.order_by('id_justificacion')
    return render(request, 'listar_justificaciones.html', {'justificaciones': justificaciones})

@login_required
def detalleJustificaciones(request, id_justificacion):
    if request.method == 'GET':
        justificacion = get_object_or_404(Justificaciones, pk=id_justificacion)
        form = FormJustificaciones(instance=justificacion)
        return render(request, 'justificacion_detalle.html', {'justificacion': justificacion, 'form': form})
    else:
        try:
            justificacion = get_object_or_404(
                Justificaciones, pk=id_justificacion)
            form = FormJustificaciones(request.POST, instance=justificacion)
            form.save()
            return redirect('justificaciones')
        except ValueError:
            return render(request, 'justificacion_detalle.html', {'justificacion': justificacion, 'form': form, 'error': 'Error al actualizar la programacion'})

# CRU Empleado (futura implementación)

# Buscar Registros
@login_required
def buscarFechaPorRango(request):
    registros = None
    if request.method == 'POST':
        fecha_inicial = request.POST.get('fecha_inicial')
        fecha_final = request.POST.get('fecha_final')

        fecha_inicial = datetime.strptime(fecha_inicial, '%Y-%m-%d').date()
        fecha_final = datetime.strptime(fecha_final, '%Y-%m-%d').date()

        registros = Registros.objects.filter(
            fecha_registro__range=[fecha_inicial, fecha_final])

    return render(request, 'buscar_fecha.html', {'registros': registros})

@login_required
def buscarEmpleado(request):
    registros = None
    if request.method == 'POST':
        id_empleado = request.POST.get('id_empleado')
        fecha_inicial = request.POST.get('fecha_inicial')
        fecha_final = request.POST.get('fecha_final')

        fecha_inicial = datetime.strptime(fecha_inicial, '%Y-%m-%d').date()
        fecha_final = datetime.strptime(fecha_final, '%Y-%m-%d').date()

        registros = Registros.objects.filter(
            empleado__id_empleado=id_empleado, fecha_registro__range=[fecha_inicial, fecha_final])

    return render(request, 'busqueda_empleado.html', {'registros': registros})

# Buscar Turnos
@login_required
def buscarTurnoFecha(request):
    turnos_programados = None
    if request.method == 'POST':
        fecha_inicial = request.POST.get('fecha_inicial')
        fecha_final = request.POST.get('fecha_final')

        fecha_inicial = datetime.strptime(fecha_inicial, '%Y-%m-%d').date()
        fecha_final = datetime.strptime(fecha_final, '%Y-%m-%d').date()

        turnos_programados = TurnosProgramados.objects.filter(
            fecha_turno__range=[fecha_inicial, fecha_final])

    return render(request, 'buscar_turnos_fecha.html', {'turnos_programados': turnos_programados})

@login_required
def buscarTurnoEmpleado(request):
    turnos_programados = None
    if request.method == 'POST':
        id_empleado = request.POST.get('id_empleado')
        fecha_inicial = request.POST.get('fecha_inicial')
        fecha_final = request.POST.get('fecha_final')

        fecha_inicial = datetime.strptime(fecha_inicial, '%Y-%m-%d').date()
        fecha_final = datetime.strptime(fecha_final, '%Y-%m-%d').date()

        turnos_programados = TurnosProgramados.objects.filter(
            empleado__id_empleado=id_empleado, fecha_turno__range=[fecha_inicial, fecha_final])

    return render(request, 'buscar_turnos_empleado.html', {'turnos_programados': turnos_programados})

# Carga de Archivos TXT
@login_required
def cargarProgramacionTurnos(request):
    if request.method == 'POST':
        form = UploadTextForm(request.POST, request.FILES)
        if form.is_valid():
            text_file = request.FILES['text_file']
            decoded_file = text_file.read().decode('utf-8')
            lines = decoded_file.split('\n')
            for line in lines:
                data = line.strip().split(';')
                turnos = TurnosProgramados(
                    empleado_id=data[0], turno_id=data[1], fecha_turno=data[2], justificacion_id=data[3])
                turnos.save()
            return render(request, 'success.html')
    else:
        form = UploadTextForm()
    return render(request, 'upload_turnos_programados.html', {'form': form})

@login_required
def cargarRegistros(request):
    if request.method == 'POST':
        form = UploadTextForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                text_file = request.FILES['text_file']
                decoded_file = text_file.read().decode('utf-8')
                lines = decoded_file.split('\n')
                for line in lines:
                    data = line.strip().split(';')
                    if len(data) == 6:
                        empleado_id = data[0]
                        turno_programado_id = data[1]
                        fecha_registro_entrada = data[2] if data[2] else None
                        hora_registro_entrada = data[3] if data[3] else None
                        fecha_registro_salida = data[4] if data[4] else None
                        hora_registro_salida = data[5] if data[5] else None

                        fecha_registro_entrada = datetime.strptime(
                            fecha_registro_entrada, "%Y-%m-%d").date() if fecha_registro_entrada else None
                        hora_registro_entrada = datetime.strptime(
                            hora_registro_entrada, "%H:%M:%S").time() if hora_registro_entrada else None
                        fecha_registro_salida = datetime.strptime(
                            fecha_registro_salida, "%Y-%m-%d").date() if fecha_registro_salida else None
                        hora_registro_salida = datetime.strptime(
                            hora_registro_salida, "%H:%M:%S").time() if hora_registro_salida else None

                        registros = Registros(empleado_id=empleado_id, turno_programado_id=turno_programado_id, fecha_registro_entrada=fecha_registro_entrada,
                                              hora_registro_entrada=hora_registro_entrada, fecha_registro_salida=fecha_registro_salida, hora_registro_salida=hora_registro_salida)
                        registros.save()
                return render(request, 'success.html')
            except Exception as e:
                return HttpResponse(f'Error al procesar el archivo: {str(e)}')
    else:
        form = UploadTextForm()
    return render(request, 'upload_registros.html', {'form': form})

# EXPORTAR INFORMACION
# Exportar datos filtrados por rango de fecha a Excel
@login_required
def exportarDatosRangoFechaExcel(request):
    fecha_inicial = request.GET.get('fecha_inicial')
    fecha_final = request.GET.get('fecha_final')

    registros = Registros.objects.filter(
        fecha_registro__range=[fecha_inicial, fecha_final])

    # Libro Excel
    workbook = Workbook()
    sheet = workbook.active

    headers = ['documento', 'id_turno_programado', 'turno_ingreso', 'fecha_entrada', 'hora_entrada',
               'fecha_salida', 'hora_salida', 'justificacion', 'marcacion', 'entrada', 'salida', 'analisis']
    sheet.append(headers)

    for registro in registros:
        row_data = [registro.empleado.id_empleado, registro.turno_programado.id_turno_programado, registro.turno.id_turno, registro.fecha_registro_entrada, registro.hora_registro_entrada, registro.fecha_registro_salida,
                    registro.hora_registro_salida, registro.turno_programado.justificacion.descripcion_justificacion, registro.obs_marcacion, registro.obs_entrada, registro.obs_salida, registro.analisis]
        sheet.append(row_data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=registros_fecha.xlsx'
    workbook.save(response)

    return response

# Exportar datos filtrados por colaborador y fecha especifica a Excel
@login_required
def exportarDatosEmpFechaExcel(request):
    id_empleado = request.GET.get('id_empleado')
    fecha_inicial = request.GET.get('fecha_inicial')
    fecha_final = request.GET.get('fecha_final')

    registros = Registros.objects.filter(
        empleado__id_empleado=id_empleado, fecha_registro__range=[fecha_inicial, fecha_final])

    # Libro Excel
    workbook = Workbook()
    sheet = workbook.active

    headers = ['documento', 'id_turno_programado', 'turno_ingreso', 'fecha_entrada', 'hora_entrada',
               'fecha_salida', 'hora_salida', 'justificacion', 'marcacion', 'entrada', 'salida', 'analisis']
    sheet.append(headers)

    for registro in registros:
        row_data = [registro.empleado.id_empleado, registro.turno_programado.id_turno_programado, registro.turno.id_turno, registro.fecha_registro_entrada, registro.hora_registro_entrada, registro.fecha_registro_salida,
                    registro.hora_registro_salida, registro.turno_programado.justificacion.descripcion_justificacion, registro.obs_marcacion, registro.obs_entrada, registro.obs_salida, registro.analisis]
        sheet.append(row_data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=registros_empleado_fecha.xlsx'
    workbook.save(response)

    return response

# Exportar turnos filtrados por rango de fecha a Excel
@login_required
def exportarTurnosRangoFechaExcel(request):
    fecha_inicial = request.GET.get('fecha_inicial')
    fecha_final = request.GET.get('fecha_final')

    turnos = TurnosProgramados.objects.filter(
        fecha_turno__range=[fecha_inicial, fecha_final])

    # Libro Excel
    workbook = Workbook()
    sheet = workbook.active

    headers = ['id_turno', 'documento', 'nombre', 'primer_apellido', 'segundo_apellido',
               'turno_asignado', 'hora_inicio_turno', 'hora_fin_turno', 'fecha_turno_asignado', 'justificacion']
    sheet.append(headers)

    for turno in turnos:
        row_data = [turno.id_turno_programado, turno.empleado.id_empleado, turno.empleado.nombre_empleado, turno.empleado.apellido_paterno_empleado,
                    turno.empleado.apellido_materno_empleado, turno.turno.id_turno, turno.turno.hora_inicial, turno.turno.hora_final, turno.fecha_turno, turno.justificacion.descripcion_justificacion]
        sheet.append(row_data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=turnos_fecha.xlsx'
    workbook.save(response)

    return response

# Exportar datos filtrados por colaborador y fecha especifica a Excel
@login_required
def exportarTurnosEmpFechaExcel(request):
    id_empleado = request.GET.get('id_empleado')
    fecha_inicial = request.GET.get('fecha_inicial')
    fecha_final = request.GET.get('fecha_final')

    turnos = TurnosProgramados.objects.filter(
        empleado__id_empleado=id_empleado, fecha_turno__range=[fecha_inicial, fecha_final])

    # Libro Excel
    workbook = Workbook()
    sheet = workbook.active

    headers = ['id_turno', 'documento', 'nombre', 'primer_apellido', 'segundo_apellido',
               'turno_asignado', 'hora_inicio_turno', 'hora_fin_turno', 'fecha_turno_asignado', 'justificacion']
    sheet.append(headers)

    for turno in turnos:
        row_data = [turno.id_turno_programado, turno.empleado.id_empleado, turno.empleado.nombre_empleado, turno.empleado.apellido_paterno_empleado,
                    turno.empleado.apellido_materno_empleado, turno.turno.id_turno, turno.turno.hora_inicial, turno.turno.hora_final, turno.fecha_turno, turno.justificacion.descripcion_justificacion]
        sheet.append(row_data)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=turnos_empleado_fecha.xlsx'
    workbook.save(response)

    return response

#Permite obtener los turnos programados de un empleado en especifico
@login_required
def get_turnos_programados(_request, empleado_id):
    try:
        #empleado_id = request.GET.get('empleado')
        turnos_programados = TurnosProgramados.objects.filter(
            empleado_id=empleado_id)
        if (len(turnos_programados) > 0):
            data = {
                turno.id_turno_programado: turno.fecha_turno for turno in turnos_programados}
        else:
            data = {'message': 'NO DATA'}
    except ValueError:
        data = {'message': 'Invalid empleado_id'}

    return JsonResponse(data)
